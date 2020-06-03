# This is a programme for tweet archive in js file translating into excel file and improve its visibility
# This programme works on MacOS

import os
import openpyxl

# get absolute path of text file
original_file = input("text file name:")
original_file_path = os.path.abspath(original_file)

# set input file name
excel_name = input("excel file name:")
new_excel_path = os.path.abspath(excel_name)

# make new excel file
wb = openpyxl.Workbook() 
sheet = wb.active

# define variables
count = 0

year = ""
month = ""
date = ""
day = ""
time = ""

tweet = ""
mention = ""
pic = ""

max_length_year = 0
max_length_month = 0
max_length_date = 0
max_length_day = 0
max_length_time = 0
max_length_mention = 0
max_length_tweet = 0
max_length_pic = 0

for line in open(original_file_path, "r", encoding = "utf_8"):
    if 'created_at' in line:
        count += 1
        year = line[-7:-3]
        month = line[24:27]
        date = line[28:30]
        day = line[20:23]
        time = line[31:39]

    if 'full_text' in line:
        tweet = line[19:-3]
    if 'in_reply_to_screen_name' in line:
        mention = "@" + line[33:-3]
    if '"media_url_https"' in line:
        pic = line[29:-3]
    if '"tweet"' in line:
        if count == 0:
            continue
        else:
            # input data into cells
            sheet.cell(row=count, column=1).value = int(year)
            sheet.cell(row=count, column=2).value = month
            sheet.cell(row=count, column=3).value = int(date)
            sheet.cell(row=count, column=4).value = day
            sheet.cell(row=count, column=5).value = time
            sheet.cell(row=count, column=6).value = mention
            sheet.cell(row=count, column=7).value = tweet
            sheet.cell(row=count, column=8).value = pic

            # set column width

            if max_length_year < len(year):
                max_length_year = len(year)
                sheet.column_dimensions['A'].width = max_length_year + 2

            if max_length_month < len(month):
                max_length_month = len(month)
                sheet.column_dimensions['B'].width = max_length_month + 2
            
            if max_length_date < len(date):
                max_length_date = len(date)
                sheet.column_dimensions['C'].width = max_length_date + 2
            
            if max_length_day < len(day):
                max_length_day = len(day)
                sheet.column_dimensions['D'].width = max_length_day + 2
            
            if max_length_time < len(time):
                max_length_time = len(time)
                sheet.column_dimensions['E'].width = max_length_time + 2

            if max_length_mention < len(mention):
                max_length_mention = len(mention)
                sheet.column_dimensions['F'].width = max_length_mention + 2
            
            if max_length_tweet < len(tweet):
                max_length_tweet = len(tweet)
                sheet.column_dimensions['G'].width = max_length_tweet + 2
             
            if max_length_pic < len(pic):
                max_length_pic = len(pic)
                sheet.column_dimensions['H'].width = max_length_pic + 2
            
            # reset some virables for next loop
            mention = ""
            pic = ""


# save and close excel file
wb.save(new_excel_path)
wb.close()

# End message
print("DONE! GO CHECK THE NEW FILE")